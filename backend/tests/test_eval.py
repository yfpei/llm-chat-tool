import pytest
from unittest.mock import patch
import openpyxl
import os
import tempfile

from app.services.eval_service import run_classification_eval


class TestClassificationEval:
    def test_basic_metrics(self):
        """Test classification metrics with simple data."""
        tmpdir = tempfile.mkdtemp()
        file_id = "test_eval_001"
        result_path = os.path.join(tmpdir, file_id + "_result.xlsx")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="input")
        ws.cell(row=1, column=2, value="label")
        ws.cell(row=1, column=3, value="AI结果")
        data = [
            ("q1", "1", "正面"),
            ("q2", "1", "正面"),
            ("q3", "1", "负面"),
            ("q4", "1", "正面"),
            ("q5", "0", "负面"),
            ("q6", "0", "负面"),
            ("q7", "0", "正面"),
            ("q8", "0", "负面"),
            ("q9", "1", "正面"),
            ("q10", "0", "负面"),
        ]
        for i, (inp, lbl, pred) in enumerate(data, start=2):
            ws.cell(row=i, column=1, value=inp)
            ws.cell(row=i, column=2, value=lbl)
            ws.cell(row=i, column=3, value=pred)
        wb.save(result_path)
        wb.close()

        with patch("app.services.eval_service.UPLOAD_DIR", tmpdir):
            result = run_classification_eval(
                file_id,
                label_column="label",
                predict_column="AI结果",
                mappings=[
                    {"model_output": "正面", "label_value": "1"},
                    {"model_output": "负面", "label_value": "0"},
                ],
            )

        assert result["total_samples"] == 10
        assert result["num_classes"] == 2
        assert result["accuracy"] == 0.8
        assert len(result["per_class"]) == 2
        assert "micro_avg" in result
        assert "macro_avg" in result
        assert result["skipped_count"] == 0
        assert os.path.exists(os.path.join(tmpdir, file_id + "_eval_classification.xlsx"))

    def test_missing_columns(self):
        tmpdir = tempfile.mkdtemp()
        file_id = "test_eval_002"
        result_path = os.path.join(tmpdir, file_id + "_result.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="col_a")
        wb.save(result_path)
        wb.close()

        with patch("app.services.eval_service.UPLOAD_DIR", tmpdir):
            with pytest.raises(ValueError, match="Columns not found"):
                run_classification_eval(
                    file_id,
                    label_column="label",
                    predict_column="AI结果",
                    mappings=[],
                )

    def test_no_result_file(self):
        with pytest.raises(FileNotFoundError):
            run_classification_eval(
                "nonexistent",
                label_column="label",
                predict_column="AI结果",
                mappings=[],
            )
