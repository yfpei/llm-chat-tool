import json
import os
import shutil
import uuid

import openpyxl
from fastapi import APIRouter, Body, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sse_starlette.sse import EventSourceResponse

from app.database import get_db
from app.deps import get_current_user
from app.models import BatchTask, User
from app.schemas import BatchUploadResponse, BatchRunRequest, FilterConfig
from app.services import batch_service

router = APIRouter(prefix="/api/batch", tags=["batch"])

ALLOWED_EXTENSIONS = (".xlsx", ".xls", ".json", ".txt")


def _text_to_excel(file_path: str, info: dict):
    """Write parsed text data into an xlsx workbook and save to file_path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    columns = info["columns"]
    # Write header row
    for ci, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=ci, value=col_name)
    # Write data rows
    for ri, row_data in enumerate(info["rows"], start=2):
        for ci, col_name in enumerate(columns, start=1):
            ws.cell(row=ri, column=ci, value=row_data.get(col_name, ""))
    # Remove the 'rows' key before passing to response (not JSON-serializable in the response)
    wb.save(file_path)
    wb.close()


@router.post("/upload")
async def upload_excel(
    file: UploadFile = File(...),
    task_id: str = Form(default=""),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not file.filename or not file.filename.lower().endswith(ALLOWED_EXTENSIONS):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls / .json / .txt 文件")

    batch_service.ensure_upload_dir()
    file_id = str(uuid.uuid4())
    ext = os.path.splitext(file.filename.lower())[1]
    is_text = ext in (".json", ".txt")
    xlsx_path = os.path.join(batch_service.UPLOAD_DIR, f"{file_id}.xlsx")

    content = await file.read()

    if is_text:
        # Save raw text temporarily, parse, then convert to xlsx
        text = content.decode("utf-8")
        tmp_path = os.path.join(batch_service.UPLOAD_DIR, f"{file_id}_raw.txt")
        with open(tmp_path, "w", encoding="utf-8") as f:
            f.write(text)
        try:
            info = batch_service.parse_text_upload(tmp_path)
        except Exception as e:
            os.remove(tmp_path)
            raise HTTPException(status_code=400, detail=f"无法解析文件: {e}")
        # Convert to Excel and remove internal rows key
        _text_to_excel(xlsx_path, info)
        os.remove(tmp_path)
        info.pop("rows", None)
    else:
        with open(xlsx_path, "wb") as f:
            f.write(content)
        try:
            info = batch_service.parse_upload(xlsx_path)
        except Exception as e:
            os.remove(xlsx_path)
            raise HTTPException(status_code=400, detail=f"无法解析 Excel 文件: {e}")

    # Keep a pristine copy so repeated batch runs always start from the original
    shutil.copy(xlsx_path, xlsx_path.replace(".xlsx", "_original.xlsx"))

    if task_id:
        # Re-upload: update existing task
        task = await db.get(BatchTask, task_id)
        if not task:
            os.remove(xlsx_path)
            original = xlsx_path.replace(".xlsx", "_original.xlsx")
            if os.path.exists(original):
                os.remove(original)
            raise HTTPException(status_code=404, detail="任务不存在")
        if task.user_id != current_user.id:
            os.remove(xlsx_path)
            original = xlsx_path.replace(".xlsx", "_original.xlsx")
            if os.path.exists(original):
                os.remove(original)
            raise HTTPException(status_code=404, detail="任务不存在")
        # Remove old files
        old_path = os.path.join(batch_service.UPLOAD_DIR, f"{task.file_id}.xlsx")
        old_original = old_path.replace(".xlsx", "_original.xlsx")
        for p in (old_path, old_original):
            if os.path.exists(p):
                os.remove(p)
        task.file_id = file_id
        task.filename = file.filename or task.filename
        task.columns = json.dumps(info["columns"], ensure_ascii=False)
        task.headers = json.dumps(info["headers"], ensure_ascii=False)
        task.total_rows = info["total_rows"]
        task.status = "uploaded"
        task.progress_completed = 0
        task.progress_total = 0
        task.config_json = None
        await db.commit()
        await db.refresh(task)
        return BatchUploadResponse(task_id=task.id, file_id=file_id, filename=file.filename or task.filename, **info)

    title = os.path.splitext(file.filename)[0] if file.filename else "未命名任务"
    task = BatchTask(
        title=title,
        file_id=file_id,
        filename=file.filename or "unknown.txt",
        columns=json.dumps(info["columns"], ensure_ascii=False),
        headers=json.dumps(info["headers"], ensure_ascii=False),
        total_rows=info["total_rows"],
        user_id=current_user.id,
    )
    db.add(task)
    await db.commit()
    await db.refresh(task)

    return BatchUploadResponse(task_id=task.id, file_id=file_id, filename=file.filename, **info)


@router.post("/run")
async def run_batch(
    req: BatchRunRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    task = await db.get(BatchTask, req.task_id)
    if not task or task.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="任务不存在")

    async def event_generator():
        async for event in batch_service.run_batch(
            task_id=req.task_id,
            file_id=req.file_id,
            input_columns=req.input_columns,
            output_column_name=req.output_column_name,
            prompt_template=req.prompt,
            api_key_id=req.api_key_id,
            concurrency=req.concurrency,
            db=db,
            strip_thinking=req.strip_thinking,
            parse_json=req.parse_json,
            filter_config=req.filter.model_dump() if req.filter else None,
            user_id=current_user.id,
        ):
            yield {
                "event": "message",
                "data": json.dumps(event, ensure_ascii=False),
            }

    return EventSourceResponse(event_generator())


@router.post("/{task_id}/filter-preview")
async def filter_preview_endpoint(
    task_id: str,
    filter_config: FilterConfig | None = Body(None),
    current_user: User = Depends(get_current_user),
):
    from app.database import async_session
    async with async_session() as db:
        task = await db.get(BatchTask, task_id)
        if not task or task.user_id != current_user.id:
            raise HTTPException(status_code=404, detail="任务不存在")
        file_id = task.file_id

    try:
        return batch_service.filter_preview(
            file_id=file_id,
            filter_config=filter_config.model_dump() if filter_config else None,
        )
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.post("/{task_id}/filter-download")
async def download_filtered(
    task_id: str,
    filter_config: FilterConfig | None = Body(None),
    current_user: User = Depends(get_current_user),
):
    from app.database import async_session
    async with async_session() as db:
        task = await db.get(BatchTask, task_id)
        if not task or task.user_id != current_user.id:
            raise HTTPException(status_code=404, detail="任务不存在")
        file_id = task.file_id
        filename = task.filename
        task_title = task.title

    original_path = os.path.join(batch_service.UPLOAD_DIR, f"{file_id}_original.xlsx")
    file_path = original_path if os.path.exists(original_path) else os.path.join(batch_service.UPLOAD_DIR, f"{file_id}.xlsx")
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="原始文件不存在")

    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active
    header_row = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []))
    headers = [str(c) if c is not None else "" for c in header_row]

    inputs: list[tuple[int, dict[str, str]]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data: dict[str, str] = {}
        for ci, val in enumerate(row):
            if ci < len(headers) and val is not None:
                row_data[headers[ci]] = str(val)
        inputs.append((row_idx, row_data))
    wb.close()

    fc = filter_config.model_dump() if filter_config else None
    filtered = batch_service.apply_filters(inputs, fc)

    # Build output workbook with only filtered rows
    out_path = os.path.join(batch_service.UPLOAD_DIR, f"{file_id}_filtered.xlsx")
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    for ci, col_name in enumerate(headers, start=1):
        ws_out.cell(row=1, column=ci, value=col_name)
    for out_row, (_, row_data) in enumerate(filtered, start=2):
        for ci, col_name in enumerate(headers, start=1):
            ws_out.cell(row=out_row, column=ci, value=row_data.get(col_name, ""))
    wb_out.save(out_path)
    wb_out.close()

    name = task_title if task_title else os.path.splitext(filename)[0]
    download_name = f"{name}_筛选结果.xlsx"
    return FileResponse(
        path=out_path,
        filename=download_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/{task_id}/download")
async def download_result(
    task_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    task = await db.get(BatchTask, task_id)
    if not task or task.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="任务不存在")
    result_path = os.path.join(batch_service.UPLOAD_DIR, f"{task.file_id}_result.xlsx")
    file_path = os.path.join(batch_service.UPLOAD_DIR, f"{task.file_id}.xlsx")
    if os.path.exists(result_path):
        file_path = result_path
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="结果文件不存在")
    name = task.title if task.title else os.path.splitext(task.filename)[0]
    download_name = f"{name}_结果.xlsx"
    return FileResponse(
        path=file_path,
        filename=download_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
