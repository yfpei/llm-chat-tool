import json
import os
import re
import asyncio
from typing import AsyncGenerator

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import ApiKey
from app.services.key_service import get_decrypted_key
from app.services.llm import create_provider

UPLOAD_DIR = os.environ.get(
    "UPLOAD_DIR",
    os.path.join(os.path.dirname(__file__), "..", "..", "uploads"),
)


def _confusion_matrix(y_true: list[str], y_pred: list[str], labels: list[str]) -> list[list[int]]:
    """Pure Python equivalent of sklearn.metrics.confusion_matrix."""
    n = len(labels)
    label_to_idx = {label: i for i, label in enumerate(labels)}
    cm = [[0] * n for _ in range(n)]
    for t, p in zip(y_true, y_pred):
        cm[label_to_idx[t]][label_to_idx[p]] += 1
    return cm


def _accuracy_score(y_true: list[str], y_pred: list[str]) -> float:
    """Pure Python equivalent of sklearn.metrics.accuracy_score."""
    if not y_true:
        return 0.0
    return sum(1 for t, p in zip(y_true, y_pred) if t == p) / len(y_true)


def _find_result_file(file_id: str) -> str | None:
    """Find result or uploaded xlsx file for a given file_id.

    Prefers _result.xlsx (batch output), falls back to .xlsx (standalone eval).
    """
    base = os.path.join(UPLOAD_DIR, file_id)
    result_path = base + "_result.xlsx"
    if os.path.exists(result_path):
        return result_path
    upload_path = base + ".xlsx"
    if os.path.exists(upload_path):
        return upload_path
    return None


def run_classification_eval(
    file_id: str,
    label_column: str,
    predict_column: str,
    mappings: list[dict],  # [{model_output, label_value}]
) -> dict:
    """Calculate classification metrics from the batch result file.

    Returns a dict matching ClassificationEvalResponse schema.
    """
    result_path = _find_result_file(file_id)
    if not result_path:
        raise FileNotFoundError(f"Result file not found for {file_id}")

    mapping_dict: dict[str, str] = {m["model_output"]: m["label_value"] for m in mappings}

    wb = openpyxl.load_workbook(result_path)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if label_column not in headers or predict_column not in headers:
        wb.close()
        missing = []
        if label_column not in headers:
            missing.append(label_column)
        if predict_column not in headers:
            missing.append(predict_column)
        raise ValueError(f"Columns not found: {', '.join(missing)}")

    label_idx = headers.index(label_column)
    predict_idx = headers.index(predict_column)

    y_true: list[str] = []
    y_pred: list[str] = []
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        label_val = str(row[label_idx]).strip() if row[label_idx] is not None else ""
        raw_predict = str(row[predict_idx]).strip() if row[predict_idx] is not None else ""

        # Apply mapping
        mapped = mapping_dict.get(raw_predict, raw_predict)

        if not label_val or not mapped:
            skipped += 1
            continue

        y_true.append(label_val)
        y_pred.append(mapped)

    wb.close()

    if not y_true:
        raise ValueError("No valid data rows found after applying mapping")

    labels = sorted(set(y_true + y_pred))
    cm = _confusion_matrix(y_true, y_pred, labels)
    cm_list: list[list[int]] = cm

    accuracy = _accuracy_score(y_true, y_pred)

    # Per-class metrics
    per_class: list[dict] = []
    per_class_raw: list[tuple[float, float, float]] = []
    for i, label in enumerate(labels):
        tp = cm[i][i]
        fp = sum(cm[j][i] for j in range(len(labels))) - tp
        fn = sum(cm[i][j] for j in range(len(labels))) - tp

        precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
        recall = tp / (tp + fn) if (tp + fn) > 0 else 0.0
        f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0.0

        per_class_raw.append((precision, recall, f1))
        per_class.append({
            "class_name": label,
            "precision": round(precision, 4) if (tp + fp) > 0 else None,
            "recall": round(recall, 4) if (tp + fn) > 0 else None,
            "f1": round(f1, 4) if (precision + recall) > 0 else None,
        })

    # Micro average: single-label → micro == accuracy
    micro_precision = accuracy
    micro_recall = accuracy
    micro_f1 = accuracy

    # Macro average: unweighted mean of per-class metrics
    n_classes = len(labels)
    if n_classes:
        macro_precision = sum(p for p, _, _ in per_class_raw) / n_classes
        macro_recall = sum(r for _, r, _ in per_class_raw) / n_classes
        macro_f1 = sum(f for _, _, f in per_class_raw) / n_classes
    else:
        macro_precision = macro_recall = macro_f1 = 0.0

    result = {
        "accuracy": round(accuracy, 4),
        "total_samples": len(y_true),
        "num_classes": len(labels),
        "confusion_matrix": cm_list,
        "labels": labels,
        "per_class": per_class,
        "micro_avg": {
            "precision": round(micro_precision, 4),
            "recall": round(micro_recall, 4),
            "f1": round(micro_f1, 4),
        },
        "macro_avg": {
            "precision": round(macro_precision, 4),
            "recall": round(macro_recall, 4),
            "f1": round(macro_f1, 4),
        },
        "skipped_count": skipped,
    }

    # Write result Excel
    _write_classification_excel(
        file_id, labels, cm_list, accuracy,
        per_class, micro_precision, micro_recall, micro_f1,
        macro_precision, macro_recall, macro_f1,
    )

    # Save result JSON for later retrieval (page refresh)
    json_path = os.path.join(UPLOAD_DIR, file_id + "_eval_classification.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False)

    return result


def _write_classification_excel(
    file_id: str, labels: list[str], cm_list: list[list[int]],
    accuracy: float, per_class: list[dict],
    micro_p: float, micro_r: float, micro_f1: float,
    macro_p: float, macro_r: float, macro_f1: float,
):
    """Write classification eval results to _eval_classification.xlsx."""
    out_path = os.path.join(UPLOAD_DIR, file_id + "_eval_classification.xlsx")
    wb = openpyxl.Workbook()

    # Sheet 1: Confusion Matrix
    ws1 = wb.active
    ws1.title = "混淆矩阵"
    # Clear row/column labels: 真实值 = actual (rows), 预测值 = predicted (columns)
    ws1.cell(row=1, column=1, value="真实值 \\ 预测值")
    for j, label in enumerate(labels, start=2):
        ws1.cell(row=1, column=j, value=label)
    col_sum_col = len(labels) + 2
    ws1.cell(row=1, column=col_sum_col, value="合计")
    for i, label in enumerate(labels):
        ws1.cell(row=i + 2, column=1, value=label)
        row_sum = 0
        for j in range(len(labels)):
            ws1.cell(row=i + 2, column=j + 2, value=cm_list[i][j])
            row_sum += cm_list[i][j]
        ws1.cell(row=i + 2, column=col_sum_col, value=row_sum)
    row_sum_row = len(labels) + 2
    ws1.cell(row=row_sum_row, column=1, value="合计")
    total_all = 0
    for j in range(len(labels)):
        col_sum = sum(cm_list[i][j] for i in range(len(labels)))
        ws1.cell(row=row_sum_row, column=j + 2, value=col_sum)
        total_all += col_sum
    ws1.cell(row=row_sum_row, column=col_sum_col, value=total_all)

    # Sheet 2: Per-class metrics
    ws2 = wb.create_sheet("各分类指标")
    ws2.cell(row=1, column=1, value="分类")
    ws2.cell(row=1, column=2, value="Precision")
    ws2.cell(row=1, column=3, value="Recall")
    ws2.cell(row=1, column=4, value="F1 Score")
    for i, pc in enumerate(per_class, start=2):
        ws2.cell(row=i, column=1, value=pc["class_name"])
        ws2.cell(row=i, column=2, value=pc["precision"] if pc["precision"] is not None else "-")
        ws2.cell(row=i, column=3, value=pc["recall"] if pc["recall"] is not None else "-")
        ws2.cell(row=i, column=4, value=pc["f1"] if pc["f1"] is not None else "-")

    # Sheet 3: Summary metrics
    ws3 = wb.create_sheet("汇总指标")
    ws3.cell(row=1, column=1, value="指标")
    ws3.cell(row=1, column=2, value="Precision")
    ws3.cell(row=1, column=3, value="Recall")
    ws3.cell(row=1, column=4, value="F1 Score")
    ws3.cell(row=2, column=1, value="Accuracy")
    ws3.cell(row=2, column=2, value=round(accuracy, 4))
    ws3.cell(row=3, column=1, value="Micro Avg")
    ws3.cell(row=3, column=2, value=round(micro_p, 4))
    ws3.cell(row=3, column=3, value=round(micro_r, 4))
    ws3.cell(row=3, column=4, value=round(micro_f1, 4))
    ws3.cell(row=4, column=1, value="Macro Avg")
    ws3.cell(row=4, column=2, value=round(macro_p, 4))
    ws3.cell(row=4, column=3, value=round(macro_r, 4))
    ws3.cell(row=4, column=4, value=round(macro_f1, 4))

    wb.save(out_path)
    wb.close()


def get_classification_eval_file(file_id: str) -> str | None:
    path = os.path.join(UPLOAD_DIR, file_id + "_eval_classification.xlsx")
    if os.path.exists(path):
        return path

    # If Excel doesn't exist but JSON does, regenerate Excel from JSON
    json_path = os.path.join(UPLOAD_DIR, file_id + "_eval_classification.json")
    if os.path.exists(json_path):
        with open(json_path, "r", encoding="utf-8") as f:
            result = json.load(f)
        # Regenerate Excel from saved result
        _write_classification_excel(
            file_id, result["labels"], result["confusion_matrix"],
            result["accuracy"], result["per_class"],
            result["micro_avg"]["precision"], result["micro_avg"]["recall"], result["micro_avg"]["f1"],
            result["macro_avg"]["precision"], result["macro_avg"]["recall"], result["macro_avg"]["f1"],
        )
        return path

    return None


def get_classification_result_json(file_id: str) -> dict | None:
    """Read saved classification result JSON."""
    json_path = os.path.join(UPLOAD_DIR, file_id + "_eval_classification.json")
    if not os.path.exists(json_path):
        return None
    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)


def get_llm_scoring_result(file_id: str) -> dict | None:
    """Read LLM scoring results from Excel, return {scores, avg_score, score_column}."""
    scoring_path = os.path.join(UPLOAD_DIR, file_id + "_eval_llm_scoring.xlsx")
    if not os.path.exists(scoring_path):
        return None

    wb = openpyxl.load_workbook(scoring_path)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    wb.close()

    # Re-read to get scores (last column)
    wb2 = openpyxl.load_workbook(scoring_path)
    ws2 = wb2.active
    score_col_name = headers[-1]
    scores: list[dict] = []
    numeric_scores: list[float] = []

    for row_idx, row in enumerate(ws2.iter_rows(min_row=2, values_only=True), start=2):
        score_val = str(row[-1]).strip() if len(row) > len(headers) - 1 and row[-1] is not None else ""
        if not score_val:
            continue
        scores.append({"row": row_idx, "score": score_val, "status": "success"})
        try:
            numeric_scores.append(float(score_val))
        except ValueError:
            pass

    wb2.close()
    avg_score = round(sum(numeric_scores) / len(numeric_scores), 2) if numeric_scores else 0

    return {
        "scores": scores,
        "avg_score": avg_score,
        "total": len(scores),
        "score_column": score_col_name,
    }


async def run_llm_scoring(
    db: AsyncSession,
    file_id: str,
    api_key_id: int,
    prompt_template: str,
    output_column_name: str,
    concurrency: int = 3,
    input_columns: list[str] | None = None,
) -> AsyncGenerator[dict, None]:
    """SSE generator for LLM subjective scoring.

    Yields: progress, row_result, row_error, done, error
    """
    result_path = _find_result_file(file_id)
    if not result_path:
        yield {"type": "error", "message": "结果文件不存在，请先跑批"}
        return

    # Get API key
    api_key = await db.get(ApiKey, api_key_id)
    if not api_key:
        yield {"type": "error", "message": "API Key 不存在"}
        return
    plaintext_key = get_decrypted_key(api_key)

    wb = openpyxl.load_workbook(result_path)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    if input_columns:
        missing = [c for c in input_columns if c not in headers]
        if missing:
            yield {"type": "error", "message": f"列 {missing} 不存在"}
            wb.close()
            return

    # Read all rows (process every row, not just those with score_column)
    rows_data: list[tuple[int, dict[str, str]]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_dict: dict[str, str] = {}
        for ci, col_name in enumerate(headers):
            if ci < len(row) and row[ci] is not None:
                row_dict[col_name] = str(row[ci])
        # Only include rows that have at least one input_column value
        if input_columns and any(row_dict.get(c) for c in input_columns):
            rows_data.append((row_idx, row_dict))
        elif not input_columns:
            # If no input_columns specified, process all rows with any data
            if row_dict:
                rows_data.append((row_idx, row_dict))
    wb.close()

    total = len(rows_data)
    if total == 0:
        yield {"type": "error", "message": "无数据行"}
        return

    yield {"type": "progress", "completed": 0, "total": total}

    # Set up scoring output file
    scoring_path = os.path.join(UPLOAD_DIR, file_id + "_eval_llm_scoring.xlsx")
    if os.path.exists(scoring_path):
        wb = openpyxl.load_workbook(scoring_path)
    else:
        wb = openpyxl.load_workbook(result_path)
    ws = wb.active
    existing_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Find or create output column
    score_out_col = None
    for col_idx, h in enumerate(existing_headers, start=1):
        if h == output_column_name:
            score_out_col = col_idx
            break
    if score_out_col is None:
        score_out_col = len(existing_headers) + 1
        ws.cell(row=1, column=score_out_col, value=output_column_name)
    wb.save(scoring_path)
    wb.close()

    results: dict[int, str] = {}
    results_lock = asyncio.Lock()
    completed = 0
    completed_lock = asyncio.Lock()
    sem = asyncio.Semaphore(concurrency)

    use_native_thinking = api_key.enable_thinking

    async def process_one(row_idx: int, row_data: dict[str, str]):
        prompt = prompt_template
        for col_name, cell_value in row_data.items():
            prompt = prompt.replace("{{" + col_name + "}}", str(cell_value))
        try:
            provider = create_provider(
                api_key.provider, api_key.base_url, plaintext_key,
                api_key.model, use_native_thinking,
            )
            messages = [{"role": "user", "content": prompt}]
            full = ""
            async for chunk in provider.chat_stream(messages):
                full += chunk
            score = full.strip()
            # Remove thinking if present
            score = re.sub(
                r'</?(?:think|unused\d+)[^>]*>.*?</(?:think|unused\d+)>',
                '', score, flags=re.DOTALL | re.IGNORECASE
            ).strip()
            async with results_lock:
                results[row_idx] = score
            return (row_idx, score, None)
        except Exception as e:
            return (row_idx, "", str(e))

    async def limited_process(row_idx: int, row_data: dict[str, str]):
        async with sem:
            return await process_one(row_idx, row_data)

    tasks_list = [asyncio.create_task(limited_process(r, d)) for r, d in rows_data]
    pending = set(tasks_list)

    while pending:
        done, pending = await asyncio.wait(pending, return_when=asyncio.FIRST_COMPLETED)
        for t in done:
            row_idx, score, error = t.result()
            async with completed_lock:
                completed += 1
            yield {"type": "progress", "completed": completed, "total": total}
            if error:
                yield {"type": "row_error", "row": row_idx, "error": error}
            else:
                yield {"type": "row_result", "row": row_idx, "score": score}

    # Write scores to Excel
    wb = openpyxl.load_workbook(scoring_path)
    ws = wb.active
    for row_idx, score in results.items():
        ws.cell(row=row_idx, column=score_out_col, value=score)
    wb.save(scoring_path)
    wb.close()

    # Calculate average
    numeric_scores = []
    for s in results.values():
        try:
            numeric_scores.append(float(s))
        except ValueError:
            pass
    avg_score = round(sum(numeric_scores) / len(numeric_scores), 2) if numeric_scores else 0

    yield {"type": "done", "total": total, "avg_score": avg_score}


def get_llm_scoring_file(file_id: str) -> str | None:
    path = os.path.join(UPLOAD_DIR, file_id + "_eval_llm_scoring.xlsx")
    return path if os.path.exists(path) else None
