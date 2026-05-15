import json
import os
import uuid
from typing import Generator

import openpyxl

UPLOAD_DIR = os.environ.get(
    "UPLOAD_DIR",
    os.path.join(os.path.dirname(__file__), "..", "..", "uploads"),
)


def ensure_upload_dir():
    os.makedirs(UPLOAD_DIR, exist_ok=True)


def get_connection(host: str, port: int, username: str, password: str, database: str = ""):
    import pymysql
    kwargs = {"host": host, "port": port, "user": username, "password": password, "charset": "utf8mb4"}
    if database:
        kwargs["database"] = database
    return pymysql.connect(**kwargs)


def test_connection(host: str, port: int, username: str, password: str) -> dict:
    conn = None
    try:
        conn = get_connection(host, port, username, password)
        with conn.cursor() as cursor:
            cursor.execute("SELECT VERSION()")
            version = cursor.fetchone()[0]
        return {"version": version}
    finally:
        if conn:
            conn.close()


def list_databases(host: str, port: int, username: str, password: str) -> list:
    conn = None
    try:
        conn = get_connection(host, port, username, password)
        with conn.cursor() as cursor:
            cursor.execute("SHOW DATABASES")
            rows = cursor.fetchall()
        sys_dbs = {"information_schema", "mysql", "performance_schema", "sys"}
        return [r[0] for r in rows if r[0] not in sys_dbs]
    finally:
        if conn:
            conn.close()


def list_tables(host: str, port: int, username: str, password: str, database: str) -> list:
    conn = None
    try:
        conn = get_connection(host, port, username, password, database)
        with conn.cursor() as cursor:
            cursor.execute("SHOW TABLES")
            rows = cursor.fetchall()
        return [r[0] for r in rows]
    finally:
        if conn:
            conn.close()


def get_columns(host: str, port: int, username: str, password: str, database: str, table: str) -> dict:
    conn = None
    try:
        conn = get_connection(host, port, username, password, database)
        with conn.cursor() as cursor:
            cursor.execute(f"SHOW FULL COLUMNS FROM `{table}`")
            rows = cursor.fetchall()
        fields = {}
        for r in rows:
            name = r[0]
            col_type = r[1]
            fields[name] = col_type
        return fields
    finally:
        if conn:
            conn.close()


def preview_query(
    host: str, port: int, username: str, password: str,
    database: str = "", table: str = "", where_clause: str = "",
    custom_sql: str = "", output_columns: list = None,
    page: int = 1, page_size: int = 50, top_n: int = None,
) -> dict:
    conn = None
    try:
        conn = get_connection(host, port, username, password, database)
        with conn.cursor() as cursor:
            if custom_sql.strip():
                sql = custom_sql.strip()
            else:
                cols = ", ".join(f"`{c}`" for c in output_columns) if output_columns else "*"
                sql = f"SELECT {cols} FROM `{table}`"
                if where_clause.strip():
                    sql += f" WHERE {where_clause.strip()}"

            count_sql = f"SELECT COUNT(*) FROM ({sql}) AS _sub"
            cursor.execute(count_sql)
            total = cursor.fetchone()[0]

            if top_n:
                total = min(total, top_n)

            limit = min(top_n, page_size) if top_n else page_size
            offset = (page - 1) * page_size
            paginated_sql = sql + f" LIMIT {limit} OFFSET {offset}"
            cursor.execute(paginated_sql)
            column_names = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()

            rows = []
            for row in raw_rows:
                record = {}
                for i, col in enumerate(column_names):
                    val = row[i]
                    if val is None:
                        record[col] = ""
                    elif isinstance(val, (dict, list)):
                        record[col] = json.dumps(val, ensure_ascii=False, default=str)
                    elif isinstance(val, bytes):
                        record[col] = val.decode("utf-8", errors="replace")
                    else:
                        record[col] = str(val)
                rows.append(record)

            return {"total": total, "rows": rows, "fields": column_names}
    finally:
        if conn:
            conn.close()


def export_to_excel(
    host: str, port: int, username: str, password: str,
    database: str = "", table: str = "", where_clause: str = "",
    custom_sql: str = "", output_columns: list = None,
    top_n: int = None,
) -> Generator[dict, None, None]:
    conn = None
    try:
        conn = get_connection(host, port, username, password, database)
        with conn.cursor() as cursor:
            if custom_sql.strip():
                sql = custom_sql.strip()
            else:
                cols = ", ".join(f"`{c}`" for c in output_columns) if output_columns else "*"
                sql = f"SELECT {cols} FROM `{table}`"
                if where_clause.strip():
                    sql += f" WHERE {where_clause.strip()}"

            count_sql = f"SELECT COUNT(*) FROM ({sql}) AS _sub"
            cursor.execute(count_sql)
            total = cursor.fetchone()[0]
            if top_n:
                total = min(total, top_n)

            if total == 0:
                yield {"type": "error", "content": "未查询到数据"}
                return

            limit_sql = sql + (f" LIMIT {top_n}" if top_n else "")
            cursor.execute(limit_sql)
            column_names = [desc[0] for desc in cursor.description]

            records: list[dict] = []
            batch_size = 1000
            batch = cursor.fetchmany(batch_size)
            while batch:
                for row in batch:
                    record = {}
                    for i, col in enumerate(column_names):
                        val = row[i]
                        if val is None:
                            record[col] = ""
                        elif isinstance(val, (dict, list)):
                            record[col] = json.dumps(val, ensure_ascii=False, default=str)
                        elif isinstance(val, bytes):
                            record[col] = val.decode("utf-8", errors="replace")
                        else:
                            record[col] = str(val)
                    records.append(record)

                yield {"type": "progress", "completed": len(records), "total": total}
                batch = cursor.fetchmany(batch_size)
    finally:
        if conn:
            conn.close()

    # Write Excel
    ensure_upload_dir()
    file_id = str(uuid.uuid4())
    output_path = os.path.join(UPLOAD_DIR, f"{file_id}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MySQL导出"

    if output_columns:
        columns = list(output_columns)
    else:
        columns = list(column_names) if column_names else []

    for ci, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=ci, value=col_name)

    for ri, record in enumerate(records, start=2):
        for ci, col_name in enumerate(columns, start=1):
            ws.cell(row=ri, column=ci, value=record.get(col_name, ""))

    wb.save(output_path)
    wb.close()

    yield {"type": "done", "file_id": file_id, "count": len(records)}
