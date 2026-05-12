import json
import os
import re
import shutil
import asyncio
from typing import AsyncGenerator

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import ApiKey, BatchTask, UserKeyOverride
from app.services.key_service import get_decrypted_key
from app.services.llm import create_provider

XINGHUO_THINKING_PROMPT = (
    "你能够回答用户的各种问题，回答问题能够角度全面、表述专业、重点突出。"
    "当前是慢思考模式，请你先深入剖析给出问题的关键要点与内在逻辑，生成思考过程，"
    "再根据思考过程回答给出问题。"
    "思考过程以<unused6>开头，在结尾处用<unused7>标注结束，"
    "<unused7>后为基于思考过程的回答内容"
)

UPLOAD_DIR = os.environ.get(
    "UPLOAD_DIR",
    os.path.join(os.path.dirname(__file__), "..", "..", "uploads"),
)
_batch_tasks: dict[str, dict] = {}

# Regex patterns for thinking tags (including custom model-specific variants)
_THINK_TAG_PATTERN = re.compile(r'</?(?:think|unused\d+)\s*/?>', re.IGNORECASE)


def ensure_upload_dir():
    os.makedirs(UPLOAD_DIR, exist_ok=True)


def parse_text_upload(file_path: str) -> dict:
    """Parse a .json or .txt file. Each non-empty line is a row.

    If the first non-empty line is valid JSON (dict), its keys become columns.
    Otherwise a single 'input' column is used.
    Returns the same shape as parse_upload() plus a 'rows' key for Excel creation.
    """
    with open(file_path, "r", encoding="utf-8") as f:
        lines = [line.strip() for line in f if line.strip()]

    if not lines:
        return {"columns": ["input"], "headers": ["input"], "total_rows": 0, "preview": [], "rows": []}

    # Determine mode: check if the first line is valid JSON object
    first_obj = None
    is_json_mode = False
    try:
        parsed = json.loads(lines[0])
        if isinstance(parsed, dict):
            first_obj = parsed
            is_json_mode = True
    except json.JSONDecodeError:
        pass

    if is_json_mode:
        # Collect all unique keys across all JSON lines
        all_keys: list[str] = []
        seen_keys = set()
        for line in lines:
            try:
                obj = json.loads(line)
                if isinstance(obj, dict):
                    for k in obj:
                        if k not in seen_keys:
                            seen_keys.add(k)
                            all_keys.append(k)
            except json.JSONDecodeError:
                pass

        columns = all_keys if all_keys else list(first_obj.keys())  # fallback to first line keys
        headers = columns

        rows = []
        preview_rows = []
        for row_idx, line in enumerate(lines):
            row_data: dict[str, str] = {}
            try:
                obj = json.loads(line)
                if isinstance(obj, dict):
                    for k in columns:
                        val = obj.get(k)
                        row_data[k] = str(val) if val is not None else ""
                else:
                    row_data[columns[0]] = line
            except json.JSONDecodeError:
                row_data[columns[0]] = line

            rows.append(row_data)
            if len(preview_rows) < 10:
                cells = {}
                for i, c in enumerate(columns):
                    text = row_data.get(c, "")
                    cells[c] = text[:50] + ("..." if len(text) > 50 else "")
                preview_rows.append({"row": row_idx + 2, "cells": cells})

        return {
            "columns": columns, "headers": headers,
            "total_rows": len(rows), "preview": preview_rows, "rows": rows,
        }
    else:
        columns = ["input"]
        headers = columns
        rows = [{"input": line} for line in lines]
        preview_rows = []
        for row_idx, line in enumerate(lines):
            if len(preview_rows) < 10:
                text = line[:50] + ("..." if len(line) > 50 else "")
                preview_rows.append({"row": row_idx + 2, "cells": {"input": text}})

        return {
            "columns": columns, "headers": headers,
            "total_rows": len(rows), "preview": preview_rows, "rows": rows,
        }


def parse_upload(file_path: str) -> dict:
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(c) if c is not None else "" for c in next(rows_iter, [])]
    columns = headers

    preview_rows = []
    row_count = 0
    for row_idx, row in enumerate(rows_iter, start=2):
        row_count += 1
        if len(preview_rows) < 10:
            cells = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    text = str(val) if val is not None else ""
                    cells[columns[i]] = text[:50] + ("..." if len(text) > 50 else "")
            preview_rows.append({"row": row_idx, "cells": cells})

    wb.close()
    return {"columns": columns, "headers": headers, "total_rows": row_count, "preview": preview_rows}


def remove_thinking(text: str) -> str:
    """Remove thinking content before closing think tags like </think> or </unused7>."""
    # Find the last closing think/unused tag and remove everything up to and including it
    matches = list(_THINK_TAG_PATTERN.finditer(text))
    if not matches:
        return text
    # Find the last closing tag (starts with </)
    closing_tags = [m for m in matches if m.group().startswith('</')]
    if not closing_tags:
        return text
    last_close = closing_tags[-1]
    return text[last_close.end():].strip()


def parse_json_from_text(text: str) -> dict | None:
    """Try to extract a flat dict of string fields from JSON text.

    Strips markdown code fences (```json / ```) and parses the remainder.
    Returns None if the text is not valid JSON or the result is not a dict.
    """
    cleaned = text.strip()
    # Strip markdown code fences
    if cleaned.startswith("```"):
        # Find end of opening fence line
        newline_pos = cleaned.find("\n")
        if newline_pos != -1:
            # Find closing fence
            closing_pos = cleaned.rfind("```")
            if closing_pos > newline_pos:
                cleaned = cleaned[newline_pos + 1:closing_pos].strip()
            else:
                # Closing fence on same line, e.g. ```json {}```
                cleaned = cleaned[newline_pos + 1:].strip()
                if cleaned.endswith("```"):
                    cleaned = cleaned[:-3].strip()

    try:
        data = json.loads(cleaned)
    except json.JSONDecodeError:
        return None

    if not isinstance(data, dict):
        return None

    result: dict[str, str] = {}
    for key, value in data.items():
        key_str = str(key)
        if isinstance(value, (dict, list)):
            result[key_str] = json.dumps(value, ensure_ascii=False)
        elif value is None:
            result[key_str] = ""
        else:
            result[key_str] = str(value)
    return result


def _eval_condition(val: str, op: str, cv: str) -> bool:
    if op == "contains":
        return cv in val
    elif op == "equals":
        return val == cv
    elif op == "gt":
        try:
            return float(val) > float(cv)
        except (ValueError, TypeError):
            return False
    elif op == "lt":
        try:
            return float(val) < float(cv)
        except (ValueError, TypeError):
            return False
    elif op == "gte":
        try:
            return float(val) >= float(cv)
        except (ValueError, TypeError):
            return False
    elif op == "lte":
        try:
            return float(val) <= float(cv)
        except (ValueError, TypeError):
            return False
    elif op == "not_empty":
        return bool(val and val.strip())
    elif op == "is_empty":
        return not val or not val.strip()
    return False


def apply_filters(
    inputs: list[tuple[int, dict[str, str]]],
    filter_config: dict | None,
) -> list[tuple[int, dict[str, str]]]:
    """Apply filter conditions organised in groups. Pure function, no side effects.

    Each group combines its conditions with its own AND/OR logic.
    Groups are then combined with the top-level logic.
    Example: (A AND B) OR (C AND D) → two groups (both "and"), group_logic="or".
    """
    if not filter_config:
        return inputs

    if filter_config.get("top_n"):
        inputs = inputs[:int(filter_config["top_n"])]

    groups = filter_config.get("groups") or []
    if not groups:
        return inputs

    group_logic = filter_config.get("logic", "and")

    filtered: list[tuple[int, dict[str, str]]] = []
    for row_idx, row_data in inputs:
        group_results: list[bool] = []
        for group in groups:
            conditions = group.get("conditions") or []
            if not conditions:
                group_results.append(True)
                continue
            gl = group.get("logic", "and")
            cond_results = [
                _eval_condition(row_data.get(c["field"], ""), c["operator"], c["value"])
                for c in conditions
            ]
            group_results.append(any(cond_results) if gl == "or" else all(cond_results))

        if group_logic == "or":
            if any(group_results):
                filtered.append((row_idx, row_data))
        else:
            if all(group_results):
                filtered.append((row_idx, row_data))
    return filtered


def filter_preview(
    file_id: str,
    filter_config: dict | None = None,
    input_columns: list[str] | None = None,
) -> dict:
    """Apply filters to the full dataset and return count + top-10 preview."""
    ensure_upload_dir()
    file_path = os.path.join(UPLOAD_DIR, f"{file_id}.xlsx")
    if not os.path.exists(file_path):
        raise FileNotFoundError("文件不存在，请重新上传")

    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active
    if not ws:
        return {"total_before": 0, "total_after": 0, "preview": [], "columns": []}

    header_row = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []))
    headers = [str(c) if c is not None else "" for c in header_row]

    inputs: list[tuple[int, dict[str, str]]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data: dict[str, str] = {}
        for ci, val in enumerate(row):
            if ci < len(headers) and val is not None:
                row_data[headers[ci]] = str(val)
        if not input_columns or any(col_name in row_data for col_name in input_columns):
            inputs.append((row_idx, row_data))

    wb.close()

    total_before = len(inputs)
    filtered = apply_filters(inputs, filter_config)
    total_after = len(filtered)

    preview = []
    for row_idx, row_data in filtered[:10]:
        cells = {}
        for col_name in headers:
            text = row_data.get(col_name, "")
            cells[col_name] = text[:50] + ("..." if len(text) > 50 else "")
        preview.append({"row": row_idx, "cells": cells})

    return {
        "total_before": total_before,
        "total_after": total_after,
        "preview": preview,
        "columns": headers,
    }

async def run_batch(
    task_id: str,
    file_id: str,
    input_columns: list[str],
    output_column_name: str,
    prompt_template: str,
    api_key_id: int,
    concurrency: int,
    db: AsyncSession,
    strip_thinking: bool = False,
    parse_json: bool = False,
    filter_config: dict | None = None,
    user_id: int | None = None,
) -> AsyncGenerator[dict, None]:
    ensure_upload_dir()
    file_path = os.path.join(UPLOAD_DIR, f"{file_id}.xlsx")
    if not os.path.exists(file_path):
        yield {"type": "error", "content": "文件不存在，请重新上传"}
        return

    # Restore pristine original so repeated runs don't compound
    original_path = file_path.replace(".xlsx", "_original.xlsx")
    if os.path.exists(original_path):
        shutil.copy(original_path, file_path)

    api_key = await db.get(ApiKey, api_key_id)
    if not api_key:
        yield {"type": "error", "content": "API Key 不存在"}
        return

    # Apply user overrides for shared keys
    if api_key.user_id is None and user_id is not None:
        result = await db.execute(
            select(UserKeyOverride).where(
                UserKeyOverride.user_id == user_id,
                UserKeyOverride.api_key_id == api_key.id,
            )
        )
        override = result.scalar_one_or_none()
        if override:
            if override.enable_thinking is not None:
                api_key.enable_thinking = override.enable_thinking
            if override.max_context_tokens is not None:
                api_key.max_context_tokens = override.max_context_tokens

    # Read Excel
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    header_row = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    headers = [str(c) if c is not None else "" for c in header_row]

    # Build column maps: all columns for filtering, input columns for substitution
    all_col_indices: dict[str, int] = {}
    for col_name in headers:
        all_col_indices[col_name] = headers.index(col_name)

    for col_name in input_columns:
        if col_name not in all_col_indices:
            yield {"type": "error", "content": f"列 '{col_name}' 不存在"}
            wb.close()
            return

    # Build inputs with ALL columns for filtering
    inputs: list[tuple[int, dict[str, str]]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data: dict[str, str] = {}
        for col_name, col_idx in all_col_indices.items():
            if col_idx < len(row) and row[col_idx] is not None:
                row_data[col_name] = str(row[col_idx])
        if any(col_name in row_data for col_name in input_columns):
            inputs.append((row_idx, row_data))

    # Apply row filters
    total_before = len(inputs)
    inputs = apply_filters(inputs, filter_config)
    total = len(inputs)
    if total == 0:
        if total_before > 0 and filter_config:
            yield {"type": "error", "content": "筛选后无匹配数据，请调整筛选条件"}
        else:
            yield {"type": "error", "content": "输入列无数据"}
        wb.close()
        return

    # Add output column header
    output_col_idx = len(headers) + 1
    ws.cell(row=1, column=output_col_idx, value=output_column_name)
    wb.save(file_path)
    wb.close()

    # Update task status to running
    task = await db.get(BatchTask, task_id)
    if task:
        task.status = "running"
        task.progress_completed = 0
        task.progress_total = total
        await db.commit()

    # Collect results: row_idx -> final_text
    results: dict[int, str] = {}
    # Parsed JSON fields per row (only when parse_json is enabled)
    parsed_fields: dict[int, dict[str, str]] = {}
    results_lock = asyncio.Lock()
    completed = 0
    completed_lock = asyncio.Lock()
    sem = asyncio.Semaphore(concurrency)

    plaintext_key = get_decrypted_key(api_key)

    use_native_thinking = api_key.enable_thinking
    if api_key.enable_thinking and api_key.model_type == "x1":
        use_native_thinking = False

    def build_input_label(row_data: dict[str, str]) -> str:
        return "; ".join(f"{k}: {v}" for k, v in row_data.items() if k in input_columns)

    async def process_one(row_idx: int, row_data: dict[str, str]):
        nonlocal completed
        prompt = prompt_template
        for col_name, cell_value in row_data.items():
            prompt = prompt.replace("{{" + col_name + "}}", cell_value)
        input_label = build_input_label(row_data)
        try:
            provider = create_provider(
                api_key.provider, api_key.base_url, plaintext_key,
                api_key.model, use_native_thinking,
            )
            messages = [{"role": "user", "content": prompt}]
            if api_key.enable_thinking and api_key.model_type == "x1":
                messages.insert(0, {"role": "system", "content": XINGHUO_THINKING_PROMPT})
            full = ""
            async for chunk in provider.chat_stream(messages):
                full += chunk

            # Post-process
            final_text = full
            row_parsed: dict[str, str] | None = None
            if strip_thinking:
                final_text = remove_thinking(full)
            if parse_json:
                row_parsed = parse_json_from_text(final_text)

            async with results_lock:
                results[row_idx] = final_text
                if row_parsed is not None:
                    parsed_fields[row_idx] = row_parsed

            return (row_idx, input_label, final_text, None, row_parsed)
        except Exception as e:
            return (row_idx, input_label, "", str(e), None)

    # Create coroutines limited by semaphore
    async def limited_process(row_idx: int, row_data: dict[str, str]):
        async with sem:
            return await process_one(row_idx, row_data)

    # Fire all tasks and collect results as they complete
    tasks_list = [asyncio.create_task(limited_process(r, d)) for r, d in inputs]

    pending = set(tasks_list)
    while pending:
        done, pending = await asyncio.wait(pending, return_when=asyncio.FIRST_COMPLETED)
        for t in done:
            row_idx, input_label, output, error, row_parsed = t.result()
            async with completed_lock:
                completed += 1
            yield {
                "type": "progress",
                "completed": completed,
                "total": total,
            }
            if error:
                yield {
                    "type": "row_error",
                    "row": row_idx,
                    "input": input_label,
                    "error": error,
                }
            else:
                event: dict = {
                    "type": "row_result",
                    "row": row_idx,
                    "input": input_label,
                    "output": output,
                    "success": True,
                }
                if row_parsed:
                    event["parsed"] = row_parsed
                yield event

            # Update DB progress periodically
            if completed % 5 == 0 or completed == total:
                task = await db.get(BatchTask, task_id)
                if task:
                    task.progress_completed = completed
                    await db.commit()

    # Write results to a new Excel with only filtered rows
    out_path = file_path.replace(".xlsx", "_result.xlsx")
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active

    # Determine JSON field columns if parse_json is enabled
    json_field_columns: dict[str, int] = {}
    parsed_field_names: list[str] = []
    if parse_json and parsed_fields:
        all_fields: set[str] = set()
        for fields in parsed_fields.values():
            all_fields.update(fields.keys())
        parsed_field_names = sorted(all_fields)

    # Write header row: original headers + output column + json field columns
    for ci, col_name in enumerate(headers, start=1):
        ws_out.cell(row=1, column=ci, value=col_name)
    ws_out.cell(row=1, column=len(headers) + 1, value=output_column_name)
    next_col = len(headers) + 2
    for field in parsed_field_names:
        ws_out.cell(row=1, column=next_col, value=field)
        json_field_columns[field] = next_col
        next_col += 1

    # Write data rows (only filtered/processed rows, in original row order)
    output_row = 2
    for row_idx, row_data in inputs:
        for ci, col_name in enumerate(headers, start=1):
            ws_out.cell(row=output_row, column=ci, value=row_data.get(col_name, ""))
        if row_idx in results:
            ws_out.cell(row=output_row, column=len(headers) + 1, value=results[row_idx])
            if row_idx in parsed_fields:
                for field, value in parsed_fields[row_idx].items():
                    if field in json_field_columns:
                        ws_out.cell(row=output_row, column=json_field_columns[field], value=value)
        output_row += 1

    wb_out.save(out_path)
    wb_out.close()

    # Mark task completed
    task = await db.get(BatchTask, task_id)
    if task:
        task.status = "completed"
        task.progress_completed = total
        await db.commit()

    _batch_tasks[task_id] = {"file_path": out_path, "filename": f"result_{file_id}.xlsx"}
    yield {"type": "done", "task_id": task_id}


def get_task(task_id: str) -> dict | None:
    return _batch_tasks.get(task_id)
