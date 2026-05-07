import json
import os
import uuid
from typing import Generator

import openpyxl
from elasticsearch import Elasticsearch

UPLOAD_DIR = os.environ.get(
    "UPLOAD_DIR",
    os.path.join(os.path.dirname(__file__), "..", "..", "uploads"),
)


def ensure_upload_dir():
    os.makedirs(UPLOAD_DIR, exist_ok=True)


def get_es_client(host: str, username: str = None, password: str = None) -> Elasticsearch:
    kwargs = {"hosts": [host], "verify_certs": False, "ssl_show_warn": False}
    if username and password:
        kwargs["http_auth"] = (username, password)
    return Elasticsearch(**kwargs)


def test_connection(host: str, username: str = None, password: str = None) -> dict:
    es = get_es_client(host, username, password)
    info = es.info()
    return {"version": info["version"]["number"], "cluster_name": info.get("cluster_name", "")}


def list_indices(host: str, username: str = None, password: str = None) -> list:
    es = get_es_client(host, username, password)
    result = es.cat.indices(format="json")
    return [idx["index"] for idx in result if not idx["index"].startswith(".")]


def get_mapping(host: str, username: str = None, password: str = None, index_name: str = "") -> dict:
    es = get_es_client(host, username, password)
    mapping = es.indices.get_mapping(index=index_name)
    fields = {}
    for index_key, index_data in mapping.items():
        props = index_data.get("mappings", {}).get("properties", {})
        _flatten_mapping(props, "", fields)
    return fields


def _flatten_mapping(props: dict, prefix: str, result: dict):
    for name, info in props.items():
        key = f"{prefix}.{name}" if prefix else name
        field_type = info.get("type", "object")
        if "properties" in info:
            _flatten_mapping(info["properties"], key, result)
        else:
            result[key] = field_type


def flatten_dict(d: dict, prefix: str = "") -> dict:
    result = {}
    for k, v in d.items():
        key = f"{prefix}.{k}" if prefix else k
        if isinstance(v, dict):
            result.update(flatten_dict(v, key))
        elif isinstance(v, list):
            result[key] = json.dumps(v, ensure_ascii=False)
        else:
            result[key] = str(v) if v is not None else ""
    return result


def preview_query(
    host: str, username: str = None, password: str = None,
    index_name: str = "", query_dsl: dict = None,
    output_fields: list = None, page: int = 1, page_size: int = 50,
    top_n: int = None,
) -> dict:
    es = get_es_client(host, username, password)
    body = {}
    if query_dsl:
        body["query"] = query_dsl
    else:
        body["query"] = {"match_all": {}}
    if output_fields:
        body["_source"] = output_fields

    size = min(top_n, page_size) if top_n else page_size
    result = es.search(index=index_name, body=body, from_=(page - 1) * page_size, size=size)

    total_info = result["hits"]["total"]
    total = total_info["value"] if isinstance(total_info, dict) else total_info
    if top_n and total > top_n:
        total = top_n

    rows = []
    all_fields: list[str] = []
    seen_fields = set()
    for hit in result["hits"]["hits"]:
        source = hit.get("_source", {})
        flat = flatten_dict(source)
        rows.append(flat)
        for k in flat:
            if k not in seen_fields:
                seen_fields.add(k)
                all_fields.append(k)

    return {"total": total, "rows": rows, "fields": all_fields}


def export_to_excel(
    host: str, username: str = None, password: str = None,
    index_name: str = "", query_dsl: dict = None,
    output_fields: list = None, top_n: int = None,
) -> Generator[dict, None, None]:
    """Export ES query results to Excel via scroll API, yielding progress events."""

    es = get_es_client(host, username, password)
    body = {}
    if query_dsl:
        body["query"] = query_dsl
    else:
        body["query"] = {"match_all": {}}
    if output_fields:
        body["_source"] = output_fields

    # Use smaller scroll batch for limited exports
    scroll_size = min(1000, top_n) if top_n else 1000

    scroll_id = None
    try:
        page = es.search(index=index_name, body=body, scroll="2m", size=scroll_size)
        scroll_id = page["_scroll_id"]
        hits = page["hits"]["hits"]

        total_info = page["hits"]["total"]
        total = total_info["value"] if isinstance(total_info, dict) else total_info
        if top_n:
            total = min(total, top_n)

        if total == 0:
            yield {"type": "error", "content": "未查询到数据"}
            return

        records: list[dict] = []
        all_fields: list[str] = []
        seen_fields = set()

        while hits:
            for hit in hits:
                source = hit.get("_source", {})
                flat = flatten_dict(source)
                records.append(flat)
                for k in flat:
                    if k not in seen_fields:
                        seen_fields.add(k)
                        all_fields.append(k)

                if top_n and len(records) >= top_n:
                    break

            yield {"type": "progress", "completed": len(records), "total": total}

            if top_n and len(records) >= top_n:
                break

            page = es.scroll(scroll_id=scroll_id, scroll="2m")
            scroll_id = page["_scroll_id"]
            hits = page["hits"]["hits"]
    finally:
        if scroll_id:
            es.clear_scroll(scroll_id=scroll_id)

    # Trim to top_n if overshot
    if top_n and len(records) > top_n:
        records = records[:top_n]

    # Write Excel
    ensure_upload_dir()
    file_id = str(uuid.uuid4())
    output_path = os.path.join(UPLOAD_DIR, f"{file_id}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ES导出"

    if output_fields:
        columns = list(output_fields)
        for f in all_fields:
            if f not in columns:
                columns.append(f)
    else:
        columns = all_fields

    for ci, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=ci, value=col_name)

    for ri, record in enumerate(records, start=2):
        for ci, col_name in enumerate(columns, start=1):
            ws.cell(row=ri, column=ci, value=record.get(col_name, ""))

    wb.save(output_path)
    wb.close()

    yield {"type": "done", "file_id": file_id, "count": len(records)}
